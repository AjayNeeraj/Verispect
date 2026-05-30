"""Generates Verispect-Sprint-Tracker.xlsx — the 15-day war room.
Tabs: Dashboard (funnel math + goal), Daily Log (15 days), Prospects, Pipeline."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PURPLE="7C6EFF"; DARK="111827"; LT="EDE9FE"; GRAY="6B7280"; GREEN="10B981"; WHITE="FFFFFF"; AMBER="F59E0B"
wb=openpyxl.Workbook()
bold=Font(bold=True); whiteb=Font(bold=True,color=WHITE)
pf=PatternFill("solid",fgColor=PURPLE); df=PatternFill("solid",fgColor=DARK); lf=PatternFill("solid",fgColor=LT)
gf=PatternFill("solid",fgColor=GREEN)
title=Font(bold=True,size=16,color=PURPLE); sub=Font(italic=True,color=GRAY,size=9)
thin=Side(style="thin",color="D1D5DB"); bd=Border(left=thin,right=thin,top=thin,bottom=thin); ctr=Alignment(horizontal="center")

def hdr(ws,row,cols,fill=pf,font=whiteb):
    for c,v in enumerate(cols,1):
        x=ws.cell(row=row,column=c,value=v); x.fill=fill; x.font=font; x.alignment=ctr; x.border=bd

# ── DASHBOARD ──
ws=wb.active; ws.title="Dashboard"
ws["A1"]="Verispect — 15-Day Sprint War Room"; ws["A1"].font=title
ws["A2"]="Goal: 10 sales floor · 15 target · 20 stretch. Update daily."; ws["A2"].font=sub
ws.column_dimensions["A"].width=34; ws.column_dimensions["B"].width=14
for c in "CDEF": ws.column_dimensions[c].width=13

ws["A4"]="FUNNEL MATH (to 15 closes)"; ws["A4"].font=bold; ws["A4"].fill=lf
math=[("Stage","Conv → next","Needed"),
      ("Personalized touches","7% reply",1500),
      ("Replies","40% snapshot",110),
      ("Snapshots run","35% demo",123),
      ("Demos / setup calls","35% close",43),
      ("CLOSES","—",15)]
hdr(ws,5,math[0])
for i,(s,cv,n) in enumerate(math[1:],6):
    ws.cell(row=i,column=1,value=s).border=bd
    ws.cell(row=i,column=2,value=cv).border=bd; ws.cell(row=i,column=2).alignment=ctr
    x=ws.cell(row=i,column=3,value=n); x.border=bd; x.alignment=ctr; x.font=bold
ws.cell(row=11,column=1).font=Font(bold=True,color=GREEN)

ws["A13"]="DAILY TARGETS"; ws["A13"].font=bold; ws["A13"].fill=lf
dt=[("Touches/day",100),("Snapshots/day",8),("Demos/day",3),("Closes/day",1)]
for i,(k,v) in enumerate(dt,14):
    ws.cell(row=i,column=1,value=k).border=bd
    x=ws.cell(row=i,column=2,value=v); x.border=bd; x.alignment=ctr; x.font=bold

ws["A19"]="LIVE SCOREBOARD (pulls from Daily Log)"; ws["A19"].font=bold; ws["A19"].fill=lf
score=[("Total touches","=SUM('Daily Log'!B3:B17)"),
       ("Total replies","=SUM('Daily Log'!C3:C17)"),
       ("Total snapshots","=SUM('Daily Log'!D3:D17)"),
       ("Total demos","=SUM('Daily Log'!E3:E17)"),
       ("TOTAL CLOSES","=SUM('Daily Log'!F3:F17)"),
       ("Cash banked (€)","=SUM('Daily Log'!G3:G17)"),
       ("MRR added (€)","=SUM('Daily Log'!H3:H17)"),
       ("Founding spots left","=20-SUM('Daily Log'!F3:F17)")]
for i,(k,f) in enumerate(score,20):
    ws.cell(row=i,column=1,value=k).border=bd
    x=ws.cell(row=i,column=2,value=f); x.border=bd; x.alignment=ctr; x.font=bold
    if "€" in k: x.number_format='€#,##0'
ws.cell(row=24,column=1).font=Font(bold=True,color=GREEN)
ws.cell(row=24,column=2).fill=gf; ws.cell(row=24,column=2).font=whiteb

ws["A29"]="vs GOAL"; ws["A29"].font=bold; ws["A29"].fill=lf
ws.cell(row=30,column=1,value="Closes vs floor (10)").border=bd
ws.cell(row=30,column=2,value="=B24&\" / 10\"").border=bd; ws.cell(row=30,column=2).alignment=ctr
ws.cell(row=31,column=1,value="Closes vs target (15)").border=bd
ws.cell(row=31,column=2,value="=B24&\" / 15\"").border=bd; ws.cell(row=31,column=2).alignment=ctr

# ── DAILY LOG ──
ws2=wb.create_sheet("Daily Log")
ws2["A1"]="Daily Log — fill end of each day"; ws2["A1"].font=title
cols=["Day","Touches","Replies","Snapshots","Demos","Closes","Cash €","MRR €","Top win / blocker"]
hdr(ws2,2,cols)
w=[7,10,9,11,8,8,10,9,40]
for i,wd in enumerate(w,1): ws2.column_dimensions[get_column_letter(i)].width=wd
for d in range(1,16):
    r=d+2
    ws2.cell(row=r,column=1,value=d).alignment=ctr; ws2.cell(row=r,column=1).border=bd
    for c in range(2,9):
        x=ws2.cell(row=r,column=c,value=0); x.border=bd; x.alignment=ctr
        if c in (7,8): x.number_format='€#,##0'
    ws2.cell(row=r,column=9).border=bd
# totals row
ws2.cell(row=18,column=1,value="TOTAL").font=bold
for c in range(2,9):
    L=get_column_letter(c)
    x=ws2.cell(row=18,column=c,value=f"=SUM({L}3:{L}17)"); x.font=bold; x.border=bd; x.alignment=ctr
    if c in (7,8): x.number_format='€#,##0'
for c in range(1,10): ws2.cell(row=18,column=c).fill=lf

# ── PROSPECTS ──
ws3=wb.create_sheet("Prospects")
ws3["A1"]="Prospects — the list (build 300+ Day 0)"; ws3["A1"].font=title
pcols=["Tier","Company","Person","Title","Geo","Vertical","LLM use case (personalization)","Source","LinkedIn/Email","Status","Last touch","Next step + date"]
hdr(ws3,2,pcols)
pw=[6,20,18,18,8,14,32,14,26,14,12,22]
for i,wd in enumerate(pw,1): ws3.column_dimensions[get_column_letter(i)].width=wd
# status legend + a few example rows blank
for r in range(3,40):
    for c in range(1,13):
        ws3.cell(row=r,column=c).border=bd
ws3.cell(row=42,column=1,value="Status values: New · Touched · Replied · Snapshot · Demo · WON · Lost · Nurture").font=sub
ws3.cell(row=43,column=1,value="Tier A = high-risk vertical + recent raise + EU (hit first). B = good fit. C = backup/Tier-2.").font=sub

# ── PIPELINE ──
ws4=wb.create_sheet("Pipeline")
ws4["A1"]="Pipeline — active deals"; ws4["A1"].font=title
dcols=["Company","Contact","Stage","Plan (annual/monthly)","Value €","Close prob %","Weighted €","Next step","Date"]
hdr(ws4,2,dcols)
dw=[20,18,16,18,11,12,12,26,12]
for i,wd in enumerate(dw,1): ws4.column_dimensions[get_column_letter(i)].width=wd
for r in range(3,30):
    for c in range(1,10): ws4.cell(row=r,column=c).border=bd
    ws4.cell(row=r,column=7,value=f"=IF(AND(E{r}<>\"\",F{r}<>\"\"),E{r}*F{r}/100,\"\")")
    ws4.cell(row=r,column=5).number_format='€#,##0'
    ws4.cell(row=r,column=7).number_format='€#,##0'
ws4.cell(row=31,column=1,value="Weighted pipeline:").font=bold
x=ws4.cell(row=31,column=7,value="=SUM(G3:G30)"); x.font=bold; x.number_format='€#,##0'; x.fill=lf
ws4.cell(row=32,column=1,value="Stages: Snapshot → Demo → Verbal → Payment sent → WON").font=sub

wb.save("Verispect-Sprint-Tracker.xlsx")
print("Saved Verispect-Sprint-Tracker.xlsx")
