"""
Generates Verispect-Financial-Model.xlsx — a founder's 24-month operating model.
Live formulas so assumptions flow through. Run: python build_financial_model.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter

PURPLE = "7C6EFF"
DARK = "111827"
LT = "EDE9FE"
GRAY = "6B7280"
GREEN = "10B981"
WHITE = "FFFFFF"

wb = openpyxl.Workbook()

bold = Font(bold=True)
white_bold = Font(bold=True, color=WHITE)
purple_fill = PatternFill("solid", fgColor=PURPLE)
dark_fill = PatternFill("solid", fgColor=DARK)
lt_fill = PatternFill("solid", fgColor=LT)
title_font = Font(bold=True, size=16, color=PURPLE)
sub_font = Font(italic=True, color=GRAY, size=9)
thin = Side(style="thin", color="D1D5DB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center")
money = '€#,##0'
money2 = '€#,##0.00'
pct = '0.0%'

def header(ws, row, cols, fill=purple_fill, font=white_bold):
    for c, val in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.fill = fill; cell.font = font; cell.alignment = center; cell.border = border

def title(ws, text, sub=None):
    ws["A1"] = text; ws["A1"].font = title_font
    if sub:
        ws["A2"] = sub; ws["A2"].font = sub_font

# ─────────────────────────────────────────────────────────────────────
# SHEET 1: Assumptions
# ─────────────────────────────────────────────────────────────────────
ws = wb.active; ws.title = "Assumptions"
title(ws, "Verispect — Operating Model", "All drivers live here. Change these; everything downstream recalculates. v1.0 · base year 2026")
ws.column_dimensions["A"].width = 40; ws.column_dimensions["B"].width = 16; ws.column_dimensions["C"].width = 50

rows = [
    ("DRIVER", "VALUE", "NOTE"),
    ("Pricing (€/mo)", "", ""),
    ("Pro price", 99, "self-serve workhorse"),
    ("Business price", 399, "scale-up / enterprise-facing"),
    ("Enterprise avg price", 4000, "custom 2k-10k, avg modeled"),
    ("Annual discount", 0.167, "~2 months free"),
    ("New logos / month", "", ""),
    ("New Free signups / mo (start)", 40, "grows with marketing"),
    ("Free signup monthly growth", 0.12, "compounding"),
    ("Free -> Pro conversion", 0.06, "of free base / mo"),
    ("New Pro (direct outbound) / mo", 2, "founder-led"),
    ("New Business / mo (from mo 4)", 0.7, "fractional = ~2 per 3 mo"),
    ("New Enterprise / mo (from mo 6)", 0.25, "~1 per 4 mo"),
    ("Churn (monthly logo)", "", ""),
    ("Pro churn", 0.035, "compliance stickiness"),
    ("Business churn", 0.02, ""),
    ("Enterprise churn", 0.01, "annual contracts"),
    ("Expansion", "", ""),
    ("Pro -> Business upgrade / mo", 0.03, "of Pro base"),
    ("Business -> Enterprise / mo", 0.02, "of Business base"),
    ("Costs", "", ""),
    ("COGS % of revenue", 0.12, "~88-90% gross margin"),
    ("Founder + ops cost / mo", 3000, "lean early"),
    ("Marketing spend / mo", 500, "content/ads/events"),
    ("Tooling / infra fixed / mo", 300, "SaaS stack"),
]
r = 4
for row in rows:
    if row[1] == "" and row[2] == "":
        ws.cell(row=r, column=1, value=row[0]).font = Font(bold=True, color=PURPLE)
        ws.cell(row=r, column=1).fill = lt_fill
    else:
        ws.cell(row=r, column=1, value=row[0]).border = border
        c = ws.cell(row=r, column=2, value=row[1]); c.border = border; c.font = bold; c.alignment = center
        if isinstance(row[1], float) and row[1] < 1:
            c.number_format = pct
        elif isinstance(row[1], (int, float)) and row[1] >= 50:
            c.number_format = money
        ws.cell(row=r, column=3, value=row[2]).font = sub_font
    r += 1

# Named cells for formula reference
A = {  # map label -> cell
    "pro": "B6", "biz": "B7", "ent": "B8", "anndisc": "B9",
    "freestart": "B11", "freegrowth": "B12", "free2pro": "B13",
    "newpro": "B14", "newbiz": "B15", "newent": "B16",
    "prochurn": "B18", "bizchurn": "B19", "entchurn": "B20",
    "pro2biz": "B22", "biz2ent": "B23",
    "cogs": "B25", "opex": "B26", "mktg": "B27", "infra": "B28",
}

# ─────────────────────────────────────────────────────────────────────
# SHEET 2: Projection (24 months)
# ─────────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("24-Month Projection")
title(ws2, "24-Month Projection", "Customer base, MRR, ARR, P&L. Driven by Assumptions sheet.")
cols = ["Month", "Free base", "Pro custs", "Business custs", "Enterprise custs",
        "MRR (€)", "ARR (€)", "Revenue (€)", "COGS (€)", "Gross profit (€)",
        "OpEx (€)", "Net (€)", "Cum. cash (€)"]
header(ws2, 4, cols)
widths = [8,11,10,13,15,13,14,13,12,14,11,12,14]
for i, w in enumerate(widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

start = 5
for m in range(1, 25):
    row = start + m - 1
    A_ = "Assumptions!"
    prev = row - 1
    # Month
    ws2.cell(row=row, column=1, value=m)
    # Free base
    if m == 1:
        ws2.cell(row=row, column=2, value=f"={A_}{A['freestart']}")
    else:
        ws2.cell(row=row, column=2,
                 value=f"=B{prev}*(1+{A_}{A['freegrowth']})+{A_}{A['freestart']}*(1+{A_}{A['freegrowth']})^{m-1}-B{prev}*{A_}{A['free2pro']}")
    # Pro
    if m == 1:
        ws2.cell(row=row, column=3, value=f"={A_}{A['newpro']}")
    else:
        ws2.cell(row=row, column=3,
                 value=f"=C{prev}*(1-{A_}{A['prochurn']}-{A_}{A['pro2biz']})+{A_}{A['newpro']}+B{prev}*{A_}{A['free2pro']}")
    # Business (from month 4)
    if m < 4:
        ws2.cell(row=row, column=4, value=0)
    elif m == 4:
        ws2.cell(row=row, column=4, value=f"={A_}{A['newbiz']}+C{prev}*{A_}{A['pro2biz']}")
    else:
        ws2.cell(row=row, column=4,
                 value=f"=D{prev}*(1-{A_}{A['bizchurn']}-{A_}{A['biz2ent']})+{A_}{A['newbiz']}+C{prev}*{A_}{A['pro2biz']}")
    # Enterprise (from month 6)
    if m < 6:
        ws2.cell(row=row, column=5, value=0)
    elif m == 6:
        ws2.cell(row=row, column=5, value=f"={A_}{A['newent']}+D{prev}*{A_}{A['biz2ent']}")
    else:
        ws2.cell(row=row, column=5,
                 value=f"=E{prev}*(1-{A_}{A['entchurn']})+{A_}{A['newent']}+D{prev}*{A_}{A['biz2ent']}")
    # MRR
    ws2.cell(row=row, column=6,
             value=f"=C{row}*{A_}{A['pro']}+D{row}*{A_}{A['biz']}+E{row}*{A_}{A['ent']}")
    # ARR
    ws2.cell(row=row, column=7, value=f"=F{row}*12")
    # Revenue (monthly = MRR)
    ws2.cell(row=row, column=8, value=f"=F{row}")
    # COGS
    ws2.cell(row=row, column=9, value=f"=H{row}*{A_}{A['cogs']}")
    # Gross profit
    ws2.cell(row=row, column=10, value=f"=H{row}-I{row}")
    # OpEx
    ws2.cell(row=row, column=11, value=f"={A_}{A['opex']}+{A_}{A['mktg']}+{A_}{A['infra']}")
    # Net
    ws2.cell(row=row, column=12, value=f"=J{row}-K{row}")
    # Cum cash
    if m == 1:
        ws2.cell(row=row, column=13, value=f"=L{row}")
    else:
        ws2.cell(row=row, column=13, value=f"=M{prev}+L{row}")
    # formats
    for col in range(2, 6):
        ws2.cell(row=row, column=col).number_format = '#,##0'
        ws2.cell(row=row, column=col).border = border
    for col in range(6, 14):
        ws2.cell(row=row, column=col).number_format = money
        ws2.cell(row=row, column=col).border = border
    ws2.cell(row=row, column=1).border = border; ws2.cell(row=row, column=1).alignment = center

# Highlight month 12 and 24
for hl in (start+11, start+23):
    for col in range(1, 14):
        ws2.cell(row=hl, column=col).fill = lt_fill

# ─────────────────────────────────────────────────────────────────────
# SHEET 3: Unit Economics
# ─────────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Unit Economics")
title(ws3, "Unit Economics", "LTV, CAC, payback by tier")
ws3.column_dimensions["A"].width = 32
for c in "BCD":
    ws3.column_dimensions[c].width = 16
header(ws3, 4, ["Metric", "Pro", "Business", "Enterprise"])
ue = [
    ("Price / mo (€)", 99, 399, 4000),
    ("Gross margin", 0.90, 0.89, 0.88),
    ("Avg lifetime (months)", 28, 40, 48),
    ("Monthly contribution (€)", "=B6*B7", "=C6*C7", "=D6*D7"),
    ("LTV (€)", "=B6*B7*B8", "=C6*C7*C8", "=D6*D7*D8"),
    ("CAC (€)", 150, 900, 2000),
    ("LTV:CAC", "=B9/B10", "=C9/C10", "=D9/D10"),
    ("CAC payback (months)", "=B10/B9*B8", "=C10/C9*C8", "=D10/D9*D8"),
]
r = 5
for label, *vals in ue:
    ws3.cell(row=r, column=1, value=label).font = bold
    ws3.cell(row=r, column=1).border = border
    for i, v in enumerate(vals, 2):
        cell = ws3.cell(row=r, column=i, value=v); cell.border = border; cell.alignment = center
        if "margin" in label:
            cell.number_format = pct
        elif "LTV:CAC" in label or "payback" in label or "lifetime" in label:
            cell.number_format = '0.0'
        else:
            cell.number_format = money
    r += 1
ws3.cell(row=r+1, column=1, value="Rule of thumb: LTV:CAC > 3 healthy, > 5 excellent. Payback < 12 mo healthy.").font = sub_font

# ─────────────────────────────────────────────────────────────────────
# SHEET 4: Scenarios
# ─────────────────────────────────────────────────────────────────────
ws4 = wb.create_sheet("Scenarios")
title(ws4, "Scenario Summary", "Manually adjust drivers on Assumptions and record Month-24 ARR here.")
ws4.column_dimensions["A"].width = 26
for c in "BCD":
    ws4.column_dimensions[c].width = 22
header(ws4, 4, ["Scenario", "Month-12 ARR (€)", "Month-24 ARR (€)", "Comment"])
scen = [
    ("Conservative", "", "", "lower conversion/growth"),
    ("Base (current drivers)", "='24-Month Projection'!G16", "='24-Month Projection'!G28", "drivers as set"),
    ("Aggressive", "", "", "higher outbound + 1 ent/mo"),
]
r = 5
for s in scen:
    ws4.cell(row=r, column=1, value=s[0]).font = bold; ws4.cell(row=r, column=1).border = border
    for i, v in enumerate(s[1:], 2):
        cell = ws4.cell(row=r, column=i, value=v); cell.border = border
        if i in (2,3) and v:
            cell.number_format = money; cell.alignment = center
    r += 1
ws4.cell(row=r+1, column=1, value="Break-even ≈ 34 Pro OR 9 Business OR ~1 Enterprise (vs ~€3,800 fixed/mo).").font = sub_font

wb.save("Verispect-Financial-Model.xlsx")
print("Saved Verispect-Financial-Model.xlsx")
